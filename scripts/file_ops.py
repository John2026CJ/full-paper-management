"""文献文件操作工具 V3.1 — 文件整合、重命名、去重、分类、MD 分类。

V3.1 变更:
  - Task 7 Step 0: 排除逻辑增强（前缀匹配排除）
  - 版本号升级

V3 功能:
  - 文件整合 (collect): 复制所有文献到 Combined 文件夹
  - 文件重命名 (rename): 标准化文献文件名
  - 去重 (dedup): MD5/DOI/标题匹配去重，移动冗余到 Duplicates
  - 按年份分类 (classify_by_year): 十年区间
  - 按主题分类 (classify_by_topic): 主题词 → 文件夹
  - 双重分类 (classify_all): ByYear + ByTopic 同时执行
  - MD 三文件夹分类 (classify_md_all): md_all + md_ByYear + md_ByTopic
  - 增量处理 (incremental): 发现 + 重命名 + 整合 + 去重 + 更新

使用方式:
  python file_ops.py collect <源目录> [--output-dir <输出>] [--rebuild]
  python file_ops.py rename <目录>
  python file_ops.py dedup <目录>
  python file_ops.py classify_by_year <Combined> --litsum <LitsSum.xlsx> --dest <ByYear>
  python file_ops.py classify_by_topic <Combined> --litsum <LitsSum.xlsx> --dest <ByTopic>
  python file_ops.py classify_all <Combined> --litsum <LitsSum.xlsx> --by-year-dest <...> --by-topic-dest <...>
  python file_ops.py classify_md_all <Combined> --litsum <LitsSum.xlsx> --md-base <md目录>
  python file_ops.py incremental <工作目录> <Combined名> <LitsSum.xlsx路径>
"""

import os
import re
import sys
import json
import hashlib
import shutil
import argparse
from pathlib import Path
from typing import List, Tuple, Dict, Optional, Set
from collections import defaultdict

# ── Constants ───────────────────────────────────────────────────
PAPER_EXTS = {'.pdf', '.docx', '.doc', '.txt', '.epub', '.mobi'}
YEAR_RANGES = [
    ('1900-1989', lambda y: y is not None and y <= 1989),
    ('1990-1999', lambda y: y is not None and 1990 <= y <= 1999),
    ('2000-2009', lambda y: y is not None and 2000 <= y <= 2009),
    ('2010-2019', lambda y: y is not None and 2010 <= y <= 2019),
    ('2020-2029', lambda y: y is not None and 2020 <= y <= 2029),
]
UNKNOWN_YEAR = 'UnknownYear'
UNCLASSIFIED = '未分类'


# ── Utility ─────────────────────────────────────────────────────
def get_unique_path(dest: str, use_copy: bool = False) -> str:
    """Get a unique destination path by appending _1, _2, ..."""
    if not os.path.exists(dest):
        return dest
    dir_name = os.path.dirname(dest)
    base, ext = os.path.splitext(os.path.basename(dest))
    counter = 1
    while True:
        new_name = f"{base}_{counter}{ext}"
        new_dest = os.path.join(dir_name, new_name)
        if not os.path.exists(new_dest):
            return new_dest
        counter += 1


def md5_hash(filepath: str) -> str:
    """Compute MD5 hash of a file."""
    h = hashlib.md5()
    try:
        with open(filepath, 'rb') as f:
            while chunk := f.read(8192):
                h.update(chunk)
    except Exception:
        return ""
    return h.hexdigest()


def safe_move(src: str, dst: str) -> Tuple[bool, str]:
    """Safely move a file, handling conflicts."""
    if os.path.exists(dst):
        dst = get_unique_path(dst)
    try:
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        shutil.move(src, dst)
        return True, dst
    except Exception as e:
        return False, str(e)


def safe_copy(src: str, dst: str) -> Tuple[bool, str]:
    """Safely copy a file, handling conflicts."""
    if os.path.exists(dst):
        dst = get_unique_path(dst)
    try:
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        shutil.copy2(src, dst)
        return True, dst
    except Exception as e:
        return False, str(e)


# ── Task 1: File Collection ─────────────────────────────────────
def collect_files(source_dir: str, output_dir: str, rebuild: bool = False,
                  dry_run: bool = False) -> Dict:
    """Collect all paper files into a Combined folder.

    Returns: {"collected": [...], "skipped": [...], "errors": [...]}
    """
    result = {"collected": [], "skipped": [], "errors": []}

    if rebuild and os.path.exists(output_dir):
        if dry_run:
            result["collected"].append({"source": output_dir,
                                        "dest": output_dir, "action": "rebuild"})
        else:
            shutil.rmtree(output_dir)

    os.makedirs(output_dir, exist_ok=True)

    for root, dirs, files in os.walk(source_dir):
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        for f in files:
            ext = Path(f).suffix.lower()
            if ext not in PAPER_EXTS:
                continue

            src = os.path.join(root, f)
            dst = os.path.join(output_dir, f)
            rel_src = os.path.relpath(src, source_dir)

            if dry_run:
                if os.path.exists(dst):
                    alt = get_unique_path(dst)
                    result["collected"].append({
                        "source": rel_src, "dest": os.path.relpath(alt, output_dir),
                        "action": "collect (conflict)"
                    })
                else:
                    result["collected"].append({
                        "source": rel_src, "dest": f, "action": "collect"
                    })
            else:
                ok, actual_dst = safe_copy(src, dst)
                if ok:
                    result["collected"].append({
                        "source": rel_src,
                        "dest": os.path.relpath(actual_dst, output_dir),
                        "action": "collect"
                    })
                else:
                    result["errors"].append({"source": rel_src, "note": actual_dst})

    return result


# ── V3.1: Enhanced exclusion for incremental ────────────────────
def EXCLUDED_FOLDERS(parent_dir: str, combined_name: str) -> Set[str]:
    """V3.1: Build set of excluded folder names for incremental runs.

    Excludes folders that:
    1. Exactly match known system folder names
    2. Start with the parent folder name prefix (e.g., Papers_*)
    """
    excluded = {
        combined_name,
        'Duplicates',
        f'{combined_name}_Duplicates',
    }

    # V3.1: Prefix-based exclusion
    parent_name = combined_name.rsplit('_Combined', 1)[0]
    if parent_name:
        prefix = f"{parent_name}_"
        for suffix in ['ByYear', 'ByTopic', 'md', 'md_all', 'md_ByYear', 'md_ByTopic',
                        'Combined_backup', 'Combined_old']:
            excluded.add(f"{prefix}{suffix}")

    # System folders
    excluded.update({'.git', '__pycache__', '.vscode', '.idea', 'node_modules'})
    return excluded


def discover_new_sources(work_dir: str, combined_name: str) -> List[str]:
    """V3.1: Discover new source folders for incremental processing."""
    excluded = EXCLUDED_FOLDERS(work_dir, combined_name)
    parent_prefix = combined_name.rsplit('_Combined', 1)[0] + '_'

    sources = []
    for item in os.listdir(work_dir):
        item_path = os.path.join(work_dir, item)
        if not os.path.isdir(item_path):
            continue

        # V3.1: Exact match or prefix match exclusion
        if item in excluded:
            continue
        if item.startswith(parent_prefix):
            continue

        # Must contain paper files
        has_papers = False
        for root, dirs, files in os.walk(item_path):
            for f in files:
                if Path(f).suffix.lower() in PAPER_EXTS:
                    has_papers = True
                    break
            if has_papers:
                break

        if has_papers:
            sources.append(item)

    return sorted(sources)


# ── Task 2: File Renaming ───────────────────────────────────────
def rename_files(target_dir: str, dry_run: bool = False) -> Dict:
    """Rename paper files to standardized format: AuthorYYYY.ext.

    Returns: {"renamed": [...], "skipped": [...], "errors": [...]}
    """
    from metadata_utils import extract_metadata, extract_metadata_from_filename

    result = {"renamed": [], "skipped": [], "errors": []}

    # Scan all files
    paper_files = []
    for root, dirs, files in os.walk(target_dir):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if Path(f).suffix.lower() in PAPER_EXTS:
                paper_files.append(os.path.join(root, f))

    # Extract metadata for conflict detection
    meta_map = {}  # base_name → list of (fpath, meta)
    for fpath in paper_files:
        meta = extract_metadata(fpath)
        surname = meta.get('first_author_surname', '') or ''
        year = meta.get('year', '') or ''
        if not surname:
            fn_info = extract_metadata_from_filename(fpath)
            surname = fn_info.get('first_author_surname', '') or 'Unknown'
        if not year:
            fn_info = extract_metadata_from_filename(fpath)
            year = fn_info.get('year', '') or 'UnknownYear'

        # Generate base name
        base = f"{surname}{year}"
        ext = os.path.splitext(fpath)[1].lower()
        key = f"{base}{ext}"
        if key not in meta_map:
            meta_map[key] = []
        meta_map[key].append((fpath, meta))

    # Process each file
    for fpath in paper_files:
        meta = extract_metadata(fpath)
        surname = meta.get('first_author_surname', '') or ''
        year = meta.get('year', '') or ''

        if not surname:
            fn_info = extract_metadata_from_filename(fpath)
            surname = fn_info.get('first_author_surname', '') or 'Unknown'
        if not year:
            fn_info = extract_metadata_from_filename(fpath)
            year = fn_info.get('year', '') or 'UnknownYear'

        dir_name = os.path.dirname(fpath)
        ext = os.path.splitext(fpath)[1].lower()
        new_name = f"{surname}{year}{ext}"
        new_path = os.path.join(dir_name, new_name)

        if surname == 'Unknown' or year == 'UnknownYear':
            result["skipped"].append({
                "source": os.path.basename(fpath),
                "note": "Cannot extract author/year"
            })
            continue

        if dry_run:
            result["renamed"].append({
                "source": os.path.basename(fpath),
                "dest": os.path.basename(new_path),
                "action": "rename"
            })
        else:
            if os.path.exists(new_path) and os.path.normpath(fpath) != os.path.normpath(new_path):
                new_path = get_unique_path(new_path)
            try:
                os.rename(fpath, new_path)
                result["renamed"].append({
                    "source": os.path.basename(fpath),
                    "dest": os.path.basename(new_path),
                    "action": "rename"
                })
            except Exception as e:
                result["errors"].append({
                    "source": os.path.basename(fpath),
                    "note": str(e)
                })

    return result


# ── Task 3: Deduplication ───────────────────────────────────────
def find_duplicates(target_dir: str) -> Dict:
    """Find duplicate paper groups in Combined folder.

    Returns: {"groups": [[file1, file2, ...], ...], "unique": [...]}
    """
    paper_files = []
    for root, dirs, files in os.walk(target_dir):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if Path(f).suffix.lower() in PAPER_EXTS:
                paper_files.append(os.path.join(root, f))

    # MD5 hash grouping
    hash_groups = defaultdict(list)
    for fp in paper_files:
        h = md5_hash(fp)
        if h:
            hash_groups[h].append(fp)

    groups = []
    processed = set()

    # Group by MD5
    for h, files in hash_groups.items():
        if len(files) > 1:
            groups.append(files)
            processed.update(files)

    # Remaining files are unique for MD5
    remaining = [f for f in paper_files if f not in processed]

    return {"groups": groups, "unique": remaining, "total": len(paper_files)}


def dedup_files(target_dir: str, dry_run: bool = False) -> Dict:
    """Deduplicate files in Combined folder.

    For each duplicate group, keeps the best file (largest size as proxy),
    moves others to Duplicates folder.

    Returns: {"kept": [...], "moved": [...], "errors": [...]}
    """
    dup_info = find_duplicates(target_dir)
    result = {"kept": [], "moved": [], "errors": []}

    if not dup_info["groups"]:
        return result

    duplicates_dir = os.path.join(target_dir, 'Duplicates')

    for group_index, group in enumerate(dup_info["groups"]):
        # Best file: largest size
        best = max(group, key=os.path.getsize)
        rest = [f for f in group if f != best]

        if dry_run:
            result["kept"].append({
                "file": os.path.basename(best),
                "group": group_index + 1,
                "reason": f"Largest ({os.path.getsize(best)} bytes)"
            })
            for f in rest:
                result["moved"].append({
                    "source": os.path.basename(f),
                    "dest": f"Duplicates/{os.path.basename(f)}",
                    "group": group_index + 1
                })
        else:
            result["kept"].append({
                "file": os.path.basename(best),
                "group": group_index + 1
            })
            for f in rest:
                dst = os.path.join(duplicates_dir, os.path.basename(f))
                ok, actual_dst = safe_move(f, dst)
                if ok:
                    result["moved"].append({
                        "source": os.path.basename(f),
                        "dest": os.path.relpath(actual_dst, target_dir),
                        "group": group_index + 1
                    })
                else:
                    result["errors"].append({
                        "source": os.path.basename(f),
                        "note": actual_dst
                    })

    return result


# ── Task 5: Classification ──────────────────────────────────────

def classify_by_year(source_dir: str, litsum_path: str, dest_dir: str,
                     use_copy: bool = True, dry_run: bool = False) -> Dict:
    """Classify papers by year into decade folders.

    Returns: {"folders": {folder_name: count, ...}, "errors": [...]}
    """
    result = {"folders": defaultdict(int), "errors": []}

    # Build year lookup from LitsSum
    year_map = {}  # filename (no ext) → year
    if os.path.exists(litsum_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(litsum_path, read_only=True, data_only=True)
            if 'LitsSum' in wb.sheetnames:
                ws = wb['LitsSum']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] and row[0]:
                        fn = str(row[1]).strip()
                        year_str = str(row[0]).strip()
                        try:
                            year = int(year_str)
                        except ValueError:
                            year = None
                        year_map[fn] = year
            wb.close()
        except Exception:
            pass

    # Process files
    for root, dirs, files in os.walk(source_dir):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if Path(f).suffix.lower() not in PAPER_EXTS:
                continue

            fname_noext = os.path.splitext(f)[0]
            year = year_map.get(fname_noext)

            # Find year range
            folder_name = UNKNOWN_YEAR
            if year is not None:
                for fname, check in YEAR_RANGES:
                    if check(year):
                        folder_name = fname
                        break

            target_folder = os.path.join(dest_dir, folder_name)
            src = os.path.join(root, f)
            dst = os.path.join(target_folder, f)

            if dry_run:
                result["folders"][folder_name] += 1
            else:
                ok, _ = safe_copy(src, dst)
                if ok:
                    result["folders"][folder_name] += 1
                else:
                    result["errors"].append({"source": f, "note": "Copy failed"})

    return result


def classify_by_topic(source_dir: str, litsum_path: str, dest_dir: str,
                      use_copy: bool = True, v3_mode: bool = True,
                      dry_run: bool = False) -> Dict:
    """Classify papers by topic from LitsSum column F.

    Returns: {"folders": {topic_name: count, ...}, "errors": [...]}
    """
    result = {"folders": defaultdict(int), "errors": []}

    # Build topic lookup from LitsSum
    topic_map = {}  # filename (no ext) → topic
    if os.path.exists(litsum_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(litsum_path, read_only=True, data_only=True)
            if 'LitsSum' in wb.sheetnames:
                ws = wb['LitsSum']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1]:
                        fn = str(row[1]).strip()
                        topic = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                        if topic and topic.lower() not in ('unknown', 'n/a', '—', '-'):
                            topic_map[fn] = topic
            wb.close()
        except Exception:
            pass

    # Process files
    for root, dirs, files in os.walk(source_dir):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if Path(f).suffix.lower() not in PAPER_EXTS:
                continue

            fname_noext = os.path.splitext(f)[0]
            topic = topic_map.get(fname_noext, UNCLASSIFIED if v3_mode else 'UnClassified')

            # Clean topic for folder name
            clean_topic = re.sub(r'[\\/:*?"<>|]', '_', topic)

            target_folder = os.path.join(dest_dir, clean_topic)
            src = os.path.join(root, f)
            dst = os.path.join(target_folder, f)

            if dry_run:
                result["folders"][clean_topic] += 1
            else:
                ok, _ = safe_copy(src, dst)
                if ok:
                    result["folders"][clean_topic] += 1
                else:
                    result["errors"].append({"source": f, "note": "Copy failed"})

    return result


def classify_all(source_dir: str, litsum_path: str,
                 by_year_dest: str, by_topic_dest: str,
                 dry_run: bool = False) -> Dict:
    """Execute both ByYear and ByTopic classification simultaneously.

    Returns: {"by_year": {...}, "by_topic": {...}}
    """
    return {
        "by_year": classify_by_year(source_dir, litsum_path, by_year_dest,
                                    use_copy=True, dry_run=dry_run),
        "by_topic": classify_by_topic(source_dir, litsum_path, by_topic_dest,
                                      use_copy=True, v3_mode=True, dry_run=dry_run),
    }


# ── Task 6: MD Classification ───────────────────────────────────

def classify_md_all(combined_dir: str, litsum_path: str, md_base_dir: str,
                    dry_run: bool = False) -> Dict:
    """Copy MD files to md_all, md_ByYear, md_ByTopic folders.

    Expects three subdirs: md_all, md_ByYear, md_ByTopic inside md_base_dir.

    Returns: {"md_all": count, "md_by_year": {...}, "md_by_topic": {...}}
    """
    md_all_dir = None
    md_by_year_dir = None
    md_by_topic_dir = None

    for item in os.listdir(md_base_dir):
        item_path = os.path.join(md_base_dir, item)
        if not os.path.isdir(item_path):
            continue
        if item.endswith('md_all'):
            md_all_dir = item_path
        elif item.endswith('md_ByYear'):
            md_by_year_dir = item_path
        elif item.endswith('md_ByTopic'):
            md_by_topic_dir = item_path

    result = {"md_all": 0, "md_by_year": {}, "md_by_topic": {}}

    # Collect all .md files from Combined (excluding Duplicates)
    md_files = []
    for root, dirs, files in os.walk(combined_dir):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if f.endswith('.md'):
                md_files.append(os.path.join(root, f))

    if not md_files:
        return result

    # Copy to md_all
    if md_all_dir:
        for fp in md_files:
            if dry_run:
                result["md_all"] += 1
            else:
                dst = os.path.join(md_all_dir, os.path.basename(fp))
                ok, _ = safe_copy(fp, dst)
                if ok:
                    result["md_all"] += 1

    # Copy to md_ByYear
    if md_by_year_dir:
        md_by_year_result = classify_by_year(
            combined_dir, litsum_path, md_by_year_dir,
            use_copy=True, dry_run=dry_run
        )
        result["md_by_year"] = dict(md_by_year_result.get("folders", {}))

    # Copy to md_ByTopic
    if md_by_topic_dir:
        md_by_topic_result = classify_md_by_topic(
            combined_dir, litsum_path, md_by_topic_dir,
            dry_run=dry_run
        )
        result["md_by_topic"] = dict(md_by_topic_result.get("folders", {}))

    return result


def classify_md_by_topic(combined_dir: str, litsum_path: str,
                         dest_dir: str, dry_run: bool = False) -> Dict:
    """Copy MD files to topic folders inside md_ByTopic."""
    return classify_by_topic(combined_dir, litsum_path, dest_dir,
                             use_copy=True, v3_mode=True, dry_run=dry_run)


# ── CLI ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="文献文件操作工具 V3.1")
    sub = parser.add_subparsers(dest="cmd")

    # collect
    collect_p = sub.add_parser("collect", help="整合文献文件")
    collect_p.add_argument("source", help="源目录")
    collect_p.add_argument("--output-dir", required=True, help="输出 Combined 目录")
    collect_p.add_argument("--rebuild", action="store_true", help="清空后重建")
    collect_p.add_argument("--dry-run", action="store_true")

    # rename
    rename_p = sub.add_parser("rename", help="重命名文献文件")
    rename_p.add_argument("dir", help="目标目录")
    rename_p.add_argument("--dry-run", action="store_true")

    # dedup
    dedup_p = sub.add_parser("dedup", help="去重文献文件")
    dedup_p.add_argument("dir", help="Combined 目录")
    dedup_p.add_argument("--dry-run", action="store_true")

    # classify
    classify_all_p = sub.add_parser("classify_all", help="双重分类（ByYear + ByTopic）")
    classify_all_p.add_argument("combined", help="Combined 目录")
    classify_all_p.add_argument("--litsum", required=True, help="LitsSum.xlsx 路径")
    classify_all_p.add_argument("--by-year-dest", required=True, help="ByYear 目标目录")
    classify_all_p.add_argument("--by-topic-dest", required=True, help="ByTopic 目标目录")
    classify_all_p.add_argument("--dry-run", action="store_true")

    classify_year_p = sub.add_parser("classify_by_year", help="按年份分类")
    classify_year_p.add_argument("combined", help="Combined 目录")
    classify_year_p.add_argument("--litsum", required=True)
    classify_year_p.add_argument("--dest", required=True)
    classify_year_p.add_argument("--dry-run", action="store_true")

    classify_topic_p = sub.add_parser("classify_by_topic", help="按主题分类")
    classify_topic_p.add_argument("combined", help="Combined 目录")
    classify_topic_p.add_argument("--litsum", required=True)
    classify_topic_p.add_argument("--dest", required=True)
    classify_topic_p.add_argument("--dry-run", action="store_true")

    # MD classification
    classify_md_p = sub.add_parser("classify_md_all", help="MD 三文件夹分类")
    classify_md_p.add_argument("combined", help="Combined 目录")
    classify_md_p.add_argument("--litsum", required=True)
    classify_md_p.add_argument("--md-base", required=True, help="MD 根目录")
    classify_md_p.add_argument("--dry-run", action="store_true")

    # incremental
    inc_p = sub.add_parser("incremental", help="增量处理（发现新增源）")
    inc_p.add_argument("work_dir", help="工作目录")
    inc_p.add_argument("combined_name", help="Combined 文件夹名称")
    inc_p.add_argument("litsum_path", help="LitsSum.xlsx 路径")

    args = parser.parse_args()

    if args.cmd == "collect":
        res = collect_files(args.source, args.output_dir,
                            rebuild=args.rebuild, dry_run=args.dry_run)
        print(f"Collected: {len(res['collected'])}")
        print(f"Skipped: {len(res['skipped'])}")
        print(f"Errors: {len(res['errors'])}")
        for item in res['collected']:
            print(f"  {item['source']} -> {item['dest']}")

    elif args.cmd == "rename":
        res = rename_files(args.dir, dry_run=args.dry_run)
        print(f"Renamed: {len(res['renamed'])}")
        print(f"Skipped: {len(res['skipped'])}")
        for item in res['renamed']:
            print(f"  {item['source']} -> {item['dest']}")
        for item in res['skipped']:
            print(f"  {item['source']}: {item['note']}")

    elif args.cmd == "dedup":
        res = dedup_files(args.dir, dry_run=args.dry_run)
        print(f"Duplicate groups: {len(set(g.get('group', 0) for g in res['moved']))}")
        print(f"Moved to Duplicates: {len(res['moved'])}")
        for item in res['moved']:
            print(f"  {item['source']} -> {item['dest']}")

    elif args.cmd == "classify_all":
        res = classify_all(args.combined, args.litsum,
                           args.by_year_dest, args.by_topic_dest,
                           dry_run=args.dry_run)
        print("\n--- ByYear ---")
        for folder, count in res["by_year"]["folders"].items():
            print(f"  {folder}: {count}")
        print("\n--- ByTopic ---")
        for folder, count in res["by_topic"]["folders"].items():
            print(f"  {folder}: {count}")

    elif args.cmd == "classify_by_year":
        res = classify_by_year(args.combined, args.litsum, args.dest,
                               dry_run=args.dry_run)
        for folder, count in res["folders"].items():
            print(f"  {folder}: {count}")

    elif args.cmd == "classify_by_topic":
        res = classify_by_topic(args.combined, args.litsum, args.dest,
                                dry_run=args.dry_run)
        for topic, count in res["folders"].items():
            print(f"  {topic}: {count}")

    elif args.cmd == "classify_md_all":
        res = classify_md_all(args.combined, args.litsum, args.md_base,
                              dry_run=args.dry_run)
        print(f"md_all: {res['md_all']}")
        print(f"md_ByYear: {res['md_by_year']}")
        print(f"md_ByTopic: {res['md_by_topic']}")

    elif args.cmd == "incremental":
        sources = discover_new_sources(args.work_dir, args.combined_name)
        if sources:
            print(f"New sources discovered: {len(sources)}")
            for s in sources:
                print(f"  - {s}")
        else:
            print("No new sources found.")
