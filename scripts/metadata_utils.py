"""文献元数据提取工具 V3.1 — 支持 PDF/DOCX/TXT 文件的完整元数据提取。

V3.1 新增:
  - 主题词库上限（100 篇文献内 ≤ 10 个主题词）
  - 超限时强制复用已有词或归入"未分类"

V3 新增:
  - 关键词提取（最多 5 个，分号分隔）
  - APA 逐字段规则构建 + 强制自检
  - 重命名冲突降级（姓氏 → 姓氏+名字全拼）
  - 非英文姓名拼音处理
  - 主题词提取（1-2 个）+ 历史词库对齐

使用方式:
  python metadata_utils.py <file_path>           # 输出 JSON 元数据
  python metadata_utils.py --batch <dir_path>    # 批量处理目录下所有文献
  python metadata_utils.py --doi <doi_string>     # 通过 DOI 在线解析元数据
"""

import os
import re
import sys
import json
import hashlib
import argparse
import traceback
from pathlib import Path
from typing import Optional, Tuple, List, Dict, Any
from collections import Counter

# ── Constants ───────────────────────────────────────────────────
PAPER_EXTS = {'.pdf', '.docx', '.doc', '.txt', '.epub', '.mobi'}
DO_NOT_CAPITALIZE = {'a', 'an', 'the', 'and', 'but', 'or', 'nor', 'for',
                     'so', 'yet', 'at', 'by', 'in', 'of', 'on', 'to', 'up',
                     'as', 'is', 'it', 'be', 'am', 'are', 'was', 'were',
                     'with', 'from', 'into', 'onto', 'upon', 'than', 'that',
                     'via', 'per', 'de'}
STOP_WORDS = {'the', 'of', 'and', 'a', 'an', 'in', 'to', 'for', 'on', 'is',
              'are', 'was', 'were', 'be', 'been', 'being', 'have', 'has', 'had',
              'do', 'does', 'did', 'but', 'or', 'not', 'no', 'if', 'we', 'you',
              'it', 'they', 'he', 'she', 'this', 'that', 'these', 'those',
              'with', 'from', 'by', 'at', 'as', 'into', 'through', 'during',
              'before', 'after', 'above', 'below', 'between', 'under', 'again',
              'further', 'then', 'once', 'here', 'there', 'all', 'both', 'each',
              'few', 'more', 'most', 'other', 'some', 'such', 'only', 'own',
              'same', 'so', 'than', 'too', 'very', 'can', 'will', 'just',
              'should', 'now', 'also', 'using', 'based', 'used', 'may', 'et', 'al',
              'due', 'new', 'well', 'much', 'since', 'within', 'without',
              'however', 'therefore', 'thus', 'although', 'while', 'where',
              'which', 'what', 'when', 'how'}

# ── V3.1: 主题词库上限 ──────────────────────────────────────────
MAX_TOPICS_PER_100 = 10           # 100 篇文献内最多 10 个主题词
TOPIC_CAP_RATIO = MAX_TOPICS_PER_100 / 100  # 每篇文献 0.1 个主题词的预算

# ── Concept maps for subject term extraction ────────────────────
SUBJECT_CONCEPT_MAP = {
    "deep learning": "Deep Learning",
    "neural network": "Deep Learning",
    "cnn": "Deep Learning",
    "rnn": "Deep Learning",
    "lstm": "Deep Learning",
    "transformer": "Deep Learning",
    "attention mechanism": "Deep Learning",
    "backpropagation": "Deep Learning",
    "convolutional": "Deep Learning",
    "recurrent": "Deep Learning",
    "generative adversarial": "Deep Learning",
    "gan": "Deep Learning",
    "autoencoder": "Deep Learning",
    "variational autoencoder": "Deep Learning",
    "diffusion model": "Deep Learning",
    "large language model": "Deep Learning",
    "llm": "Deep Learning",
    "gpt": "Deep Learning",
    "bert": "Deep Learning",
    "pre-trained": "Deep Learning",
    "foundation model": "Deep Learning",
    "machine learning": "Machine Learning",
    "random forest": "Machine Learning",
    "svm": "Machine Learning",
    "support vector": "Machine Learning",
    "decision tree": "Machine Learning",
    "xgboost": "Machine Learning",
    "gradient boosting": "Machine Learning",
    "k-means": "Machine Learning",
    "clustering": "Machine Learning",
    "ensemble": "Machine Learning",
    "natural language processing": "Deep Learning and NLP",
    "nlp": "Deep Learning and NLP",
    "text mining": "Deep Learning and NLP",
    "sentiment analysis": "Deep Learning and NLP",
    "named entity": "Deep Learning and NLP",
    "machine translation": "Deep Learning and NLP",
    "language model": "Deep Learning and NLP",
    "tokenization": "Deep Learning and NLP",
    "computer vision": "Computer Vision",
    "image recognition": "Computer Vision",
    "object detection": "Computer Vision",
    "image segmentation": "Computer Vision",
    "image classification": "Computer Vision",
    "visual recognition": "Computer Vision",
    "facial recognition": "Computer Vision",
    "reinforcement learning": "Reinforcement Learning",
    "q-learning": "Reinforcement Learning",
    "policy gradient": "Reinforcement Learning",
    "markov decision": "Reinforcement Learning",
    "robotics": "Robotics and Control",
    "control system": "Robotics and Control",
    "robust control": "Robotics and Control",
    "pid": "Robotics and Control",
    "quantum": "Quantum Computing",
    "qubit": "Quantum Computing",
    "quantum computing": "Quantum Computing",
    "quantum algorithm": "Quantum Computing",
    "bioinformatics": "Bioinformatics",
    "genomics": "Bioinformatics",
    "protein": "Bioinformatics",
    "dna": "Bioinformatics",
    "rna": "Bioinformatics",
    "drug discovery": "Bioinformatics",
    "federated learning": "Federated Learning",
    "privacy preserving": "Federated Learning",
    "differential privacy": "Federated Learning",
    "edge computing": "Edge Computing",
    "iot": "Internet of Things",
    "internet of things": "Internet of Things",
    "blockchain": "Blockchain",
    "cybersecurity": "Cybersecurity",
    "network security": "Cybersecurity",
    "encryption": "Cybersecurity",
    "cryptography": "Cybersecurity",
    "knowledge graph": "Knowledge Graph",
    "semantic web": "Knowledge Graph",
    "ontology": "Knowledge Graph",
    "time series": "Time Series Analysis",
    "forecasting": "Time Series Analysis",
    "anomaly detection": "Time Series Analysis",
    "optimization": "Optimization",
    "operations research": "Optimization",
    "linear programming": "Optimization",
    "bayesian": "Bayesian Methods",
    "probabilistic": "Bayesian Methods",
    "graph neural network": "Graph Neural Networks",
    "gnn": "Graph Neural Networks",
    "recommendation system": "Recommendation Systems",
    "collaborative filtering": "Recommendation Systems",
}

SUBJECT_CONCEPT_MAP_CN = {
    "深度学习": "Deep Learning",
    "神经网络": "Deep Learning",
    "机器学习": "Machine Learning",
    "自然语言处理": "Deep Learning and NLP",
    "计算机视觉": "Computer Vision",
    "强化学习": "Reinforcement Learning",
    "量子计算": "Quantum Computing",
    "生物信息": "Bioinformatics",
    "联邦学习": "Federated Learning",
    "边缘计算": "Edge Computing",
    "物联网": "Internet of Things",
    "区块链": "Blockchain",
    "网络安全": "Cybersecurity",
    "知识图谱": "Knowledge Graph",
    "时间序列": "Time Series Analysis",
    "优化": "Optimization",
    "贝叶斯": "Bayesian Methods",
    "图神经网络": "Graph Neural Networks",
    "推荐系统": "Recommendation Systems",
}


# ── PDF text extraction ─────────────────────────────────────────
def extract_pdf_text(filepath: str) -> Tuple[str, bool, str]:
    """Extract text layer from PDF. Returns (text, ok, error)."""
    try:
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(filepath)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            return text.strip(), True, ""
        except ImportError:
            pass

        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(filepath)
            text = ""
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text += t + "\n"
            return text.strip(), True, ""
        except ImportError:
            pass

        try:
            import pikepdf
            pdf = pikepdf.open(filepath)
            text = ""
            for page in pdf.pages:
                if hasattr(page, 'Contents'):
                    text += str(page.Contents) + "\n"
            return text.strip(), True, ""
        except ImportError:
            pass

        return "", False, "No PDF library available (install PyMuPDF, PyPDF2, or pikepdf)"
    except Exception as e:
        return "", False, f"PDF extraction failed: {e}"


def extract_pdf_metadata(filepath: str) -> Dict[str, str]:
    """Extract PDF metadata fields."""
    meta = {}
    try:
        try:
            import fitz
            doc = fitz.open(filepath)
            meta = dict(doc.metadata)
            doc.close()
            return meta
        except ImportError:
            pass

        try:
            import PyPDF2
            reader = PyPDF2.PdfReader(filepath)
            info = reader.metadata
            if info:
                meta = {k.lstrip('/'): str(v) for k, v in info.items() if v}
            return meta
        except ImportError:
            pass
    except Exception:
        pass
    return meta


# ── DOCX text extraction ────────────────────────────────────────
def extract_docx_text(filepath: str) -> Tuple[str, bool, str]:
    """Extract text from DOCX files."""
    try:
        try:
            from docx import Document
            doc = Document(filepath)
            text = "\n".join(p.text for p in doc.paragraphs)
            return text.strip(), True, ""
        except ImportError:
            return "", False, "python-docx not installed"
    except Exception as e:
        return "", False, f"DOCX extraction failed: {e}"


# ── TXT extraction ──────────────────────────────────────────────
def extract_txt_text(filepath: str) -> Tuple[str, bool, str]:
    """Read plain text file."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.read().strip(), True, ""
    except UnicodeDecodeError:
        try:
            with open(filepath, 'r', encoding='gbk') as f:
                return f.read().strip(), True, ""
        except Exception as e:
            return "", False, f"TXT encoding error: {e}"
    except Exception as e:
        return "", False, f"TXT read failed: {e}"


# ── Unified text extraction ─────────────────────────────────────
def extract_text(filepath: str, max_chars: int = 3000) -> Tuple[str, bool, str]:
    """Extract text from any supported format. Reads up to max_chars."""
    ext = Path(filepath).suffix.lower()
    if ext == '.pdf':
        text, ok, err = extract_pdf_text(filepath)
    elif ext == '.docx':
        text, ok, err = extract_docx_text(filepath)
    elif ext == '.txt':
        text, ok, err = extract_txt_text(filepath)
    elif ext in ('.doc',):
        text, ok, err = extract_docx_text(filepath)
    elif ext in ('.epub', '.mobi'):
        return "", False, f"Unsupported format for text extraction: {ext}"
    else:
        return "", False, f"Unknown format: {ext}"

    if ok and len(text) > max_chars:
        text = text[:max_chars]
    return text, ok, err


# ── Metadata extraction from text ───────────────────────────────
def extract_author_from_text(text: str) -> Optional[str]:
    """Extract first author from academic text."""
    patterns = [
        # Standard citation: Smith, J. (2023) or Smith, J.D.
        r'([A-Z][a-z]+(?:[\-\'][A-Z][a-z]+)?),\s*[A-Z](?:\.[A-Z])?\.?\s*(?:\band\b|,)',
        # Author in header: Smith et al.
        r'([A-Z][a-z]+(?:[\-\'][A-Z][a-z]+)?)\s+et\s+al\.',
        # Chinese: 张三
        r'([\u4e00-\u9fff]{1,4})\s*(?:\(|（)',
    ]
    for pat in patterns:
        m = re.search(pat, text[:2000])
        if m:
            return m.group(1)
    return None


def extract_year_from_text(text: str) -> Optional[str]:
    """Extract publication year from text."""
    patterns = [
        r'(?:19|20)\d{2}',  # Any 4-digit year in range
    ]
    years = []
    for pat in patterns:
        matches = re.findall(pat, text[:2000])
        years.extend(int(m) for m in matches if 1900 <= int(m) <= 2030)
    if years:
        return str(max(set(years), key=years.count))
    return None


def extract_title_from_text(text: str) -> Optional[str]:
    """Extract likely title from first meaningful line(s)."""
    lines = text.strip().split('\n')
    for line in lines:
        line = line.strip()
        if len(line) > 15 and len(line) < 300:
            if not any(kw in line.lower() for kw in
                       ['abstract', 'introduction', 'copyright', 'http', 'www.',
                        'university', 'department', 'correspondence']):
                return line[:100]
    return None


def extract_keywords_from_text(text: str, max_keywords: int = 5) -> str:
    """Extract keywords from text using frequency analysis."""
    words = re.findall(r'\b[a-zA-Z]{4,}\b', text.lower())
    filtered = [w for w in words if w not in STOP_WORDS]

    # Count bigram phrases
    bigrams = [' '.join(filtered[i:i+2]) for i in range(len(filtered)-1)]
    bigram_counts = Counter(bigrams)

    candidates = []
    for phrase, count in bigram_counts.most_common(20):
        if count >= 2 and not any(c in candidates for c in phrase.split()):
            candidates.append(phrase)

    if len(candidates) < max_keywords:
        word_counts = Counter(filtered)
        for word, _ in word_counts.most_common(20):
            if not any(word in c for c in candidates):
                candidates.append(word)
            if len(candidates) >= max_keywords:
                break

    result = [kw.strip().title() for kw in candidates[:max_keywords]]
    return '; '.join(result)


def extract_doi(text: str) -> Optional[str]:
    """Extract DOI from text."""
    pattern = r'(?:doi\s*:?\s*|https?://doi\.org/)(10\.\d{4,}/[^\s<>"\']+)'
    m = re.search(pattern, text, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


# ── DOI resolution ──────────────────────────────────────────────
def resolve_doi(doi: str) -> Optional[Dict[str, Any]]:
    """Resolve DOI via Crossref API. Returns metadata dict or None."""
    import urllib.request
    import urllib.error

    url = f"https://api.crossref.org/works/{doi}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "LitManager/3.1"})
        resp = urllib.request.urlopen(req, timeout=10)
        data = json.loads(resp.read().decode('utf-8'))
        msg = data.get('message', {})

        authors = msg.get('author', [])
        author_list = [a.get('family', a.get('given', '')) for a in authors]

        return {
            'title': msg.get('title', [''])[0] if msg.get('title') else '',
            'authors': [a.get('family', '') for a in authors],
            'first_author_surname': author_list[0] if author_list else '',
            'first_author_full': msg.get('author', [{}])[0].get('given', ''),
            'year': str(msg.get('created', {}).get('date-parts', [[None]])[0][0] or
                        msg.get('published-print', {}).get('date-parts', [[None]])[0][0] or ''),
            'journal': msg.get('container-title', [''])[0] if msg.get('container-title') else '',
            'volume': msg.get('volume', ''),
            'issue': msg.get('issue', ''),
            'page': msg.get('page', ''),
            'doi': doi,
            'keywords': '; '.join(msg.get('subject', [])) if msg.get('subject') else '',
        }
    except (urllib.error.URLError, urllib.error.HTTPError, json.JSONDecodeError, Exception):
        return None


# ── APA citation builder ────────────────────────────────────────
def _format_apa_authors(author_str: str) -> str:
    """Format author string to APA style: Last, F. M."""
    if not author_str:
        return ""
    if any('\u4e00' <= c <= '\u9fff' for c in author_str):
        return author_str.strip()
    names = [n.strip() for n in re.split(r'[;,|&]|\band\b', author_str) if n.strip()]
    formatted = []
    for name in names:
        parts = name.split()
        if len(parts) >= 2:
            surname = parts[0]
            initials = '. '.join(p[0].upper() + '.' for p in parts[1:])
            formatted.append(f"{surname}, {initials}")
        else:
            formatted.append(parts[0])
    if len(formatted) == 1:
        return formatted[0]
    elif len(formatted) == 2:
        return f"{formatted[0]} & {formatted[1]}"
    else:
        return ", ".join(formatted[:-1]) + f", & {formatted[-1]}"


def build_apa_citation(meta: Dict[str, Any], retry_count: int = 0) -> Tuple[str, bool, str]:
    """Build APA 7th edition citation string. Returns (citation, ok, note)."""
    MAX_RETRIES = 3
    if retry_count >= MAX_RETRIES:
        return "", False, f"Failed APA build after {MAX_RETRIES} retries"

    author = meta.get('first_author_surname', '') or meta.get('authors', '')
    year = meta.get('year', '')
    title = meta.get('title', '')
    journal = meta.get('journal', '')
    volume = meta.get('volume', '')
    issue = meta.get('issue', '')
    page = meta.get('page', '')
    doi = meta.get('doi', '')
    is_chinese = meta.get('is_chinese', False)

    parts = []

    # Author
    if author:
        if isinstance(author, list):
            author_str = ', '.join(author)
        else:
            author_str = str(author)
        if not is_chinese:
            author_str = _format_apa_authors(author_str)
        if not author_str.endswith('.'):
            author_str += '.'
        parts.append(author_str)

    # Year
    if year:
        parts.append(f"({year}).")
    elif author:
        parts.append("(n.d.).")

    # Title
    if title:
        title = title.strip()
        if not title.endswith('.'):
            title += '.'
        # First-word capitalization only for English
        if not is_chinese and not any('\u4e00' <= c <= '\u9fff' for c in title):
            words = title.split()
            if words:
                words[0] = words[0].capitalize()
                title = ' '.join(words)
        parts.append(title)

    # Journal
    if journal:
        if not is_chinese:
            journal_formatted = f"*{journal}*"
        else:
            journal_formatted = f"*{journal}*"
        # Volume, Issue
        extra = ""
        if volume:
            extra += f"*{volume}*"
        if issue:
            extra += f"({issue})"
        if extra:
            extra += ","
        if page:
            extra += f" {page}."
        journal_formatted += extra
        if not journal_formatted.endswith('.'):
            journal_formatted += '.'
        parts.append(journal_formatted)

    # DOI
    if doi:
        doi_str = f"https://doi.org/{doi}"
        if not parts[-1].endswith('.'):
            parts[-1] += f" {doi_str}"
        else:
            parts.append(doi_str)

    citation = ' '.join(parts)
    citation = re.sub(r'\s{2,}', ' ', citation)
    citation = re.sub(r'\.{2,}', '.', citation)
    citation = citation.strip()

    # Self-check
    errors = []
    if not is_chinese and author and not re.search(r'[A-Z][a-z]+,\s*[A-Z]\.', citation):
        errors.append("Author format may be incorrect")
    if year and not re.search(r'\(\d{4}\)', citation):
        errors.append("Year format incorrect")
    if doi and 'doi.org' not in citation:
        errors.append("DOI prefix missing")
    if citation.count('..') > 0:
        errors.append("Double period detected")

    if errors:
        return build_apa_citation(meta, retry_count + 1)

    return citation, True, ""


# ── Subject term extraction (V3.1 with topic cap) ───────────────
def _validate_subject_terms(terms: List[str]) -> List[str]:
    """Validate and clean subject terms."""
    cleaned = []
    special_chars = re.compile(r'[\\/:*?"<>|]')
    for term in terms:
        term = term.strip()
        term = special_chars.sub('_', term)
        # Capitalize properly
        words = term.split()
        cap_words = []
        for w in words:
            if w.upper() == w and len(w) <= 5:  # Acronyms
                cap_words.append(w.upper())
            elif w.lower() in DO_NOT_CAPITALIZE:
                cap_words.append(w.lower())
            else:
                cap_words.append(w.capitalize())
        cleaned.append(' '.join(cap_words))
    return cleaned


def _map_concept(concept: str) -> Optional[str]:
    """Map a concept keyword to a standardized subject category."""
    cl = concept.lower().strip()
    # Exact match
    if cl in SUBJECT_CONCEPT_MAP:
        return SUBJECT_CONCEPT_MAP[cl]
    if cl in SUBJECT_CONCEPT_MAP_CN:
        return SUBJECT_CONCEPT_MAP_CN[cl]
    # Substring match
    for key, val in SUBJECT_CONCEPT_MAP.items():
        if key in cl or cl in key:
            return val
    for key, val in SUBJECT_CONCEPT_MAP_CN.items():
        if key in cl or cl in key:
            return val
    return None


def extract_subject_terms(keywords_text: str, title: str = "",
                          historical_subjects: Optional[List[str]] = None,
                          total_paper_count: int = 1) -> List[str]:
    """Extract 1-2 subject terms with V3.1 topic cap enforcement.

    Args:
        keywords_text: Keywords string (semicolon-separated)
        title: Paper title
        historical_subjects: List of existing subject terms in the library
        total_paper_count: Total number of papers in the library

    Returns:
        List of 1-2 subject terms (may be empty if unclassified under cap)
    """
    if historical_subjects is None:
        historical_subjects = []

    # Collect candidate concepts
    candidates = []
    if keywords_text:
        for kw in keywords_text.split(';'):
            kw = kw.strip()
            if kw and len(kw) >= 3:
                candidates.append(kw)
    if title and not candidates:
        words = re.findall(r'\b[a-zA-Z]{4,}\b', title.lower())
        valid = [w for w in words if w not in STOP_WORDS]
        if valid:
            candidates.extend(valid[:5])

    # Map candidates to standardized concepts
    mapped_concepts = []
    seen = set()
    for c in candidates:
        mapped = _map_concept(c)
        if mapped and mapped not in seen:
            mapped_concepts.append(mapped)
            seen.add(mapped)

    # V3.1: Enforce topic word cap
    max_topics = max(1, int(total_paper_count * TOPIC_CAP_RATIO))
    current_topic_count = len(set(historical_subjects))

    # First try: align with historical subjects
    if historical_subjects:
        for concept in mapped_concepts:
            if concept in historical_subjects:
                return [concept]  # Reuse existing

    # If at cap, try to reuse closest existing topic
    if current_topic_count >= max_topics:
        if historical_subjects:
            # Find the semantically closest historical topic
            for hist_topic in historical_subjects:
                for concept in mapped_concepts:
                    if any(w.lower() in concept.lower() for w in hist_topic.split()
                           if len(w) > 3):
                        return [hist_topic]  # Reuse closest
            # No match found → unclassified
            return []
        # No historical topics but at cap (shouldn't happen)
        return []

    # Below cap → generate new topic
    if mapped_concepts:
        result = mapped_concepts[:2]
        # Two-word combine check
        if len(result) == 2:
            combined = f"{result[0]} and {result[1]}"
            if len(combined) <= 30:
                return [combined]
            else:
                return [result[0]]
        return result

    # Fallback: try direct from keywords
    if candidates:
        raw = candidates[0]
        raw = raw.strip().title()
        if len(raw) <= 30:
            return [raw]

    return []  # Unclassifiable


def load_historical_subjects(litsum_path: str) -> List[str]:
    """Load existing subject terms from LitsSum.xlsx column F."""
    subjects = []
    if not os.path.exists(litsum_path):
        return subjects
    try:
        import openpyxl
        wb = openpyxl.load_workbook(litsum_path, read_only=True, data_only=True)
        if 'LitsSum' in wb.sheetnames:
            ws = wb['LitsSum']
            for row in ws.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
                val = row[0]
                if val and str(val).strip():
                    v = str(val).strip()
                    if v.lower() not in ('unknown', 'n/a', '—', '-'):
                        subjects.append(v)
        wb.close()
    except Exception:
        pass
    return list(dict.fromkeys(subjects))  # deduplicate preserving order


# ── Filename-based metadata extraction ──────────────────────────
def extract_metadata_from_filename(filename: str) -> Dict[str, str]:
    """Extract author/year from filename patterns."""
    basename = os.path.splitext(os.path.basename(filename))[0]
    info = {}

    # Pattern: Smith2020, ZhangYuan2023, Chen2024a
    m = re.match(r'^([A-Z][a-z]+(?:[A-Z][a-z]+)?)(\d{4})$', basename)
    if m:
        info['first_author_surname'] = m.group(1)
        info['year'] = m.group(2)
    else:
        # Fallback: extract year from any position
        m = re.search(r'(\d{4})', basename)
        if m:
            y = m.group(1)
            if 1900 <= int(y) <= 2030:
                info['year'] = y
    return info


# ── Main metadata extraction ────────────────────────────────────
def extract_metadata(filepath: str, litsum_path: str = None,
                     total_papers: int = 1) -> Dict[str, Any]:
    """Extract comprehensive metadata from a literature file.

    Returns dict with keys: filepath, doi, title, first_author_surname,
    first_author_full, authors, year, journal, volume, issue, page,
    keywords, subject_terms, apa_citation, text_preview, ok, error
    """
    result = {
        'filepath': filepath,
        'doi': '',
        'title': '',
        'first_author_surname': '',
        'first_author_full': '',
        'authors': [],
        'year': '',
        'journal': '',
        'volume': '',
        'issue': '',
        'page': '',
        'keywords': '',
        'subject_terms': [],
        'apa_citation': '',
        'text_preview': '',
        'is_chinese': False,
        'ok': False,
        'error': '',
    }

    # Step 1: Extract full text
    text, ok, err = extract_text(filepath)
    if not ok:
        result['error'] = err
        return result

    result['text_preview'] = text[:500]

    # Detect language
    if any('\u4e00' <= c <= '\u9fff' for c in text):
        result['is_chinese'] = True

    # Step 2: Try DOI resolution first
    doi_match = extract_doi(text)
    doi_data = None
    if doi_match:
        result['doi'] = doi_match
        doi_data = resolve_doi(doi_match)

    if doi_data:
        result.update({k: v for k, v in doi_data.items()
                       if k in result and v})
        result['ok'] = True

    # Step 3: Extract from PDF metadata
    ext = Path(filepath).suffix.lower()
    if ext == '.pdf':
        pdf_meta = extract_pdf_metadata(filepath)
        if not result.get('title') and pdf_meta.get('title'):
            result['title'] = str(pdf_meta['title']).strip()
        if not result.get('first_author_surname'):
            author_raw = pdf_meta.get('author') or pdf_meta.get('Author', '')
            if author_raw:
                parts = re.split(r'[;,|]', str(author_raw))
                if parts:
                    surname = parts[0].strip().split()[-1]
                    result['first_author_surname'] = surname

    # Step 4: Extract from text content
    if not result.get('first_author_surname'):
        result['first_author_surname'] = extract_author_from_text(text) or ''
    if not result.get('year'):
        result['year'] = extract_year_from_text(text) or ''
    if not result.get('title'):
        result['title'] = extract_title_from_text(text) or ''

    # Keywords
    if not result.get('keywords') and not doi_data:
        result['keywords'] = extract_keywords_from_text(text)

    # Step 5: Fallback from filename
    if not result.get('first_author_surname') or not result.get('year'):
        fn_info = extract_metadata_from_filename(os.path.basename(filepath))
        if not result.get('first_author_surname'):
            result['first_author_surname'] = fn_info.get('first_author_surname', '')
        if not result.get('year'):
            result['year'] = fn_info.get('year', '')

    # Build APA
    if result.get('first_author_surname'):
        apa, apa_ok, apa_err = build_apa_citation(result)
        if apa_ok:
            result['apa_citation'] = apa
        else:
            result['apa_citation'] = f"{result['first_author_surname']} ({result.get('year', 'n.d.')}). {result.get('title', '')}"

    # Subject terms (V3.1 with cap)
    historical_subjects = []
    if litsum_path and os.path.exists(litsum_path):
        historical_subjects = load_historical_subjects(litsum_path)

    result['subject_terms'] = extract_subject_terms(
        result['keywords'],
        result['title'],
        historical_subjects,
        total_papers
    )

    result['ok'] = True
    return result


# ── Batch processing ────────────────────────────────────────────
def batch_extract(directory: str, litsum_path: str = None) -> List[Dict[str, Any]]:
    """Extract metadata from all papers in a directory."""
    results = []
    paper_files = []
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d != 'Duplicates' and not d.startswith('.')]
        for f in files:
            if Path(f).suffix.lower() in PAPER_EXTS:
                paper_files.append(os.path.join(root, f))
    paper_files.sort()

    for fpath in paper_files:
        print(f"  Processing: {os.path.basename(fpath)}", file=sys.stderr)
        meta = extract_metadata(fpath, litsum_path, total_papers=len(paper_files))
        results.append(meta)

    return results


# ── CLI ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="文献元数据提取工具 V3.1")
    sub = parser.add_subparsers(dest="cmd")

    single_p = sub.add_parser("extract", help="提取单个文件元数据")
    single_p.add_argument("path", help="文件路径")
    single_p.add_argument("--litsum", help="LitsSum.xlsx 路径（用于主题词对齐）")

    batch_p = sub.add_parser("batch", help="批量提取目录下所有文献元数据")
    batch_p.add_argument("dir", help="目录路径")
    batch_p.add_argument("--litsum", help="LitsSum.xlsx 路径")
    batch_p.add_argument("--json", action="store_true", help="输出 JSON 格式")

    doi_p = sub.add_parser("doi", help="通过 DOI 在线解析")
    doi_p.add_argument("doi_string", help="DOI 字符串")

    args = parser.parse_args()

    if args.cmd == "extract":
        meta = extract_metadata(args.path, args.litsum)
        print(json.dumps(meta, ensure_ascii=False, indent=2))

    elif args.cmd == "batch":
        results = batch_extract(args.dir, args.litsum)
        if args.json:
            print(json.dumps(results, ensure_ascii=False, indent=2))
        else:
            for r in results:
                print(f"\n{'='*60}")
                print(f"File: {os.path.basename(r['filepath'])}")
                print(f"Author: {r['first_author_surname']}")
                print(f"Year: {r['year']}")
                print(f"Title: {r['title'][:80]}")
                print(f"DOI: {r['doi']}")
                print(f"Keywords: {r['keywords']}")
                print(f"Subject: {r['subject_terms']}")
                print(f"APA: {r['apa_citation'][:120]}")
                print(f"Status: {'OK' if r['ok'] else 'FAIL: ' + r['error']}")

    elif args.cmd == "doi":
        data = resolve_doi(args.doi_string)
        if data:
            print(json.dumps(data, ensure_ascii=False, indent=2))
        else:
            print("DOI resolution failed")
