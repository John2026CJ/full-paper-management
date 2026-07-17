"""LitsSum Excel 管理工具 V3.1 — LitsSum.xlsx 的创建、读取、更新。

V3.1 无脚本变更（版本号升级）。

V3 特性:
  - 6 列结构：出版年份 | 文件名 | 完整引用(APA) | 原始文件夹 | 关键词 | 主题词
  - 增量更新：保留已有行，仅追加新行或更新已有行
  - 备份机制：写前自动备份
  - 排序：按年份升序，Unknown 排最后

使用方式:
  python litsum_manager.py create <路径>              # 创建新的 LitsSum.xlsx
  python litsum_manager.py read <路径>                # 读取并显示内容
  python litsum_manager.py update <路径> --entries <JSON>  # 更新条目
  python litsum_manager.py sort <路径>                # 排序并保存
"""

import os
import sys
import json
import shutil
import argparse
from datetime import datetime
from typing import List, Dict, Optional, Any, Tuple

# ── Constants ───────────────────────────────────────────────────
HEADERS = ['出版年份', '文件名', '完整引用(APA)', '原始文件夹', '关键词', '主题词']
NUM_COLS = 6
COL_YEAR = 0      # A
COL_FILENAME = 1  # B
COL_APA = 2       # C
COL_SOURCE = 3    # D
COL_KEYWORDS = 4  # E
COL_SUBJECT = 5   # F

# Column widths (in character units)
COL_WIDTHS = [14, 30, 80, 25, 40, 45]


# ── Backup ──────────────────────────────────────────────────────
def backup_litsum(filepath: str) -> Optional[str]:
    """Create a timestamped backup of LitsSum.xlsx."""
    if not os.path.exists(filepath):
        return None
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    dir_name = os.path.dirname(filepath)
    base = os.path.splitext(os.path.basename(filepath))[0]
    backup_name = f"{base}_backup_{timestamp}.xlsx"
    backup_path = os.path.join(dir_name, backup_name)
    try:
        shutil.copy2(filepath, backup_path)
        return backup_path
    except Exception as e:
        print(f"  Warning: Backup failed: {e}", file=sys.stderr)
        return None


# ── Create ──────────────────────────────────────────────────────
def create_litsum(filepath: str) -> bool:
    """Create a new LitsSum.xlsx with headers."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LitsSum'

        # Header styling
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Set column widths
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        wb.save(filepath)
        return True
    except ImportError:
        print("Error: openpyxl is required (pip install openpyxl)", file=sys.stderr)
        return False
    except Exception as e:
        print(f"Error creating LitsSum: {e}", file=sys.stderr)
        return False


# ── Read ────────────────────────────────────────────────────────
def read_litsum(filepath: str) -> Tuple[List[Dict[str, str]], bool]:
    """Read all rows from LitsSum.xlsx.

    Returns: (rows, ok)
    Each row is a dict with keys matching HEADERS.
    """
    if not os.path.exists(filepath):
        return [], False
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if 'LitsSum' not in wb.sheetnames:
            wb.close()
            return [], False

        ws = wb['LitsSum']
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            entry = {}
            for idx, header in enumerate(HEADERS):
                val = row[idx] if idx < len(row) else None
                entry[header] = str(val).strip() if val is not None else ''
            # Skip completely empty rows
            if any(entry[h] for h in HEADERS):
                rows.append(entry)
        wb.close()
        return rows, True
    except ImportError:
        return [], False
    except Exception as e:
        print(f"Error reading LitsSum: {e}", file=sys.stderr)
        return [], False


def get_determination_keys(rows: List[Dict]) -> List[str]:
    """Extract determination keys (filename) from rows."""
    keys = []
    for row in rows:
        fn = row.get('文件名', '')
        if fn:
            keys.append(fn)
    return keys


# ── Update ──────────────────────────────────────────────────────
def update_litsum(filepath: str, entries: List[Dict[str, str]],
                  backup: bool = True) -> Tuple[int, int, List[str]]:
    """Update LitsSum with new entries.

    Args:
        filepath: Path to LitsSum.xlsx
        entries: List of dicts with HEADERS keys
        backup: Whether to create backup before writing

    Returns:
        (added_count, updated_count, errors)
    """
    if backup:
        backup_litsum(filepath)

    # Read existing data
    existing_rows, ok = read_litsum(filepath)
    if not ok and os.path.exists(filepath):
        return 0, 0, ["Cannot read existing LitsSum"]

    # Build existing filename index (determination key)
    existing_filenames = set()
    for row in existing_rows:
        fn = row.get('文件名', '')
        if fn:
            existing_filenames.add(fn)

    added = 0
    updated = 0
    errors = []

    # Process entries
    merged = list(existing_rows)  # Start with existing

    for entry in entries:
        fn = entry.get('文件名', '')
        if not fn:
            continue

        if fn in existing_filenames:
            # Update existing row
            for idx, row in enumerate(merged):
                if row.get('文件名') == fn:
                    # Update with new data (keep more complete version)
                    if entry.get('完整引用(APA)') and (
                        not row.get('完整引用(APA)') or
                        len(entry.get('完整引用(APA)', '')) >
                        len(row.get('完整引用(APA)', ''))
                    ):
                        merged[idx]['完整引用(APA)'] = entry['完整引用(APA)']

                    # Keywords: keep the one with more entries
                    old_kw = row.get('关键词', '')
                    new_kw = entry.get('关键词', '')
                    if len(new_kw.split(';')) > len(old_kw.split(';')):
                        merged[idx]['关键词'] = new_kw

                    # Subject terms: fill empty, don't overwrite non-empty
                    if not row.get('主题词', '') and entry.get('主题词', ''):
                        merged[idx]['主题词'] = entry['主题词']

                    # Source folder: keep existing
                    if not row.get('原始文件夹', '') and entry.get('原始文件夹', ''):
                        merged[idx]['原始文件夹'] = entry['原始文件夹']

                    updated += 1
                    break
        else:
            # Add new entry
            merged.append(entry)
            existing_filenames.add(fn)
            added += 1

    # Sort by year
    def sort_key(row):
        year_str = row.get('出版年份', '')
        try:
            return (0, int(year_str))
        except (ValueError, TypeError):
            return (1, 0)  # Unknown goes last

    merged.sort(key=sort_key)

    # Write back
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'LitsSum'

        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_alignment = Alignment(vertical='top', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Headers
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        for row_idx, entry in enumerate(merged, 2):
            for col_idx, header in enumerate(HEADERS):
                value = entry.get(header, '')
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

        # Column widths
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'
        wb.save(filepath)
        return added, updated, errors
    except ImportError:
        return 0, 0, ["openpyxl is required"]
    except Exception as e:
        return 0, 0, [str(e)]


# ── Sort ────────────────────────────────────────────────────────
def sort_litsum(filepath: str) -> bool:
    """Read, sort by year, and save LitsSum."""
    rows, ok = read_litsum(filepath)
    if not ok:
        return False

    def sort_key(row):
        year_str = row.get('出版年份', '')
        try:
            return (0, int(year_str))
        except (ValueError, TypeError):
            return (1, 0)

    rows.sort(key=sort_key)

    # Convert to entries format
    entries = []
    for row in rows:
        entries.append(row)

    added, updated, errors = update_litsum(filepath, entries, backup=True)
    return len(errors) == 0


# ── CLI ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="LitsSum Excel 管理工具 V3.1")
    sub = parser.add_subparsers(dest="cmd")

    create_p = sub.add_parser("create", help="创建新的 LitsSum.xlsx")
    create_p.add_argument("path", help="保存路径")

    read_p = sub.add_parser("read", help="读取并显示 LitsSum.xlsx")
    read_p.add_argument("path", help="文件路径")

    update_p = sub.add_parser("update", help="更新 LitsSum.xlsx 条目")
    update_p.add_argument("path", help="文件路径")
    update_p.add_argument("--entries", required=True, help="JSON 格式的条目列表")
    update_p.add_argument("--no-backup", action="store_true", help="跳过备份")

    sort_p = sub.add_parser("sort", help="按年份排序")

    args = parser.parse_args()

    if args.cmd == "create":
        ok = create_litsum(args.path)
        if ok:
            print(f"Created: {args.path}")
        else:
            print("Creation failed")

    elif args.cmd == "read":
        rows, ok = read_litsum(args.path)
        if not ok:
            print(f"Cannot read: {args.path}")
            sys.exit(1)

        print(f"{'Year':<8} {'File':<30} {'Subject':<20} {'Keywords':<40}")
        print("-" * 100)
        for row in rows:
            year = row.get('出版年份', '')[:8]
            fname = row.get('文件名', '')[:28]
            subj = row.get('主题词', '')[:18]
            kw = row.get('关键词', '')[:38]
            print(f"{year:<8} {fname:<30} {subj:<20} {kw:<40}")

    elif args.cmd == "update":
        try:
            entries = json.loads(args.entries)
        except json.JSONDecodeError as e:
            print(f"Invalid JSON: {e}")
            sys.exit(1)

        added, updated, errors = update_litsum(
            args.path, entries, backup=not args.no_backup
        )
        print(f"Added: {added}, Updated: {updated}")
        if errors:
            print(f"Errors: {errors}")

    elif args.cmd == "sort":
        # Sort needs a path argument
        print("Usage: python litsum_manager.py sort <path>")
        print("Please provide the file path as first positional argument.")
